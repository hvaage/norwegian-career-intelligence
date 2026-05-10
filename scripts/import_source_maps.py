#!/usr/bin/env python3
"""
Import Spor 1 and Spor 2 source-map workbooks into Supabase:
  sources, datasets, dataset_versions

Environment (from .env or process env):
  SUPABASE_URL
  SUPABASE_SERVICE_ROLE_KEY

Workbooks (fixed paths relative to project root):
  data/source_maps/Spor1_Kildekartlegging.xlsx  -> intelligence_layer: education_supply
  data/source_maps/Spor2_Kildekartlegging.xlsx -> intelligence_layer: employer_demand
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import Client, create_client

# -----------------------------------------------------------------------------
# Paths & constants
# -----------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent.parent

SPOR1_PATH = PROJECT_ROOT / "data" / "source_maps" / "Spor1_Kildekartlegging.xlsx"
SPOR2_PATH = PROJECT_ROOT / "data" / "source_maps" / "Spor2_Kildekartlegging.xlsx"

INTELLIGENCE_LAYER_SPOR1 = "education_supply"
INTELLIGENCE_LAYER_SPOR2 = "employer_demand"

DATASET_VERSION_LABEL = "source-map-placeholder"

# Heuristic column names (normalized: lower, strip). First match wins.
NAME_KEYS = (
    "kilde",
    "kildenavn",
    "datakilde",
    "navn",
    "tittel",
    "title",
    "source",
    "dataset",
    "beskrivelse",
    "description",
    "organisasjon",
    "institusjon",
)
URL_KEYS = ("url", "lenke", "link", "nettside", "api", "api_url", "kilde_url")
KIND_KEYS = ("type", "kind", "format", "datatype", "leveranseform", "tilgang", "access")
OWNER_KEYS = ("eier", "owner", "ansvarlig", "team")
LICENSE_KEYS = ("lisens", "license", "notat", "notes", "merknad")
REFRESH_KEYS = ("frekvens", "refresh", "oppdatering", "kadence")


@dataclass
class ImportStats:
    sources_inserted: int = 0
    sources_updated: int = 0
    datasets_inserted: int = 0
    datasets_updated: int = 0
    versions_inserted: int = 0
    versions_updated: int = 0
    rows_skipped: int = 0


def _slugify(text: str, max_len: int = 80) -> str:
    s = text.strip().lower()
    s = re.sub(r"[^\w\s-]", "", s, flags=re.UNICODE)
    s = re.sub(r"[-\s]+", "-", s).strip("-")
    return s[:max_len] if s else "row"


def _normalize_header(cell: Any) -> str:
    if cell is None:
        return ""
    return str(cell).strip().lower()


def _row_to_dict(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    """Build JSON-serializable dict for metadata (dates → ISO, rest → JSON-safe types)."""
    out: dict[str, Any] = {}
    for i, key in enumerate(headers):
        if not key:
            key = f"column_{i}"
        val = row[i] if i < len(row) else None
        if val is None:
            out[key] = None
        elif hasattr(val, "isoformat") and callable(getattr(val, "isoformat")):
            out[key] = val.isoformat()
        elif isinstance(val, (int, float, bool)):
            out[key] = val
        elif isinstance(val, str):
            out[key] = val.strip()
        else:
            out[key] = str(val).strip()
    return out


def _pick_value(row_norm: dict[str, Any], candidates: tuple[str, ...]) -> str | None:
    for cand in candidates:
        if cand in row_norm and row_norm[cand] not in (None, ""):
            return str(row_norm[cand]).strip()
    for key, val in row_norm.items():
        if val in (None, ""):
            continue
        for cand in candidates:
            if cand in key.replace(" ", "_"):
                return str(val).strip()
    return None


def _content_hash(layer: str, sheet: str, row_dict: dict[str, Any]) -> str:
    payload = json.dumps(
        {"layer": layer, "sheet": sheet, "row": row_dict},
        sort_keys=True,
        ensure_ascii=False,
        default=str,
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:16]


def _stable_slug(layer: str, sheet: str, row_dict: dict[str, Any]) -> str:
    row_norm = {k.lower().strip(): v for k, v in row_dict.items() if k}
    name = _pick_value(row_norm, NAME_KEYS)
    prefix = "edu-supply" if layer == INTELLIGENCE_LAYER_SPOR1 else "emp-demand"
    h = _content_hash(layer, sheet, row_dict)
    if name:
        base = _slugify(f"{prefix}-{name}")
        if len(base) > 100:
            base = base[:100]
        return f"{base}-{h[:8]}"
    return f"{prefix}-{sheet[:40]}-{h}"


def _is_empty_row(row_dict: dict[str, Any]) -> bool:
    for v in row_dict.values():
        if v is not None and str(v).strip() != "":
            return False
    return True


def inspect_workbook(path: Path) -> None:
    print(f"\n--- Inspecting: {path} ---")
    if not path.is_file():
        print(f"  ERROR: file not found: {path}")
        return
    wb = load_workbook(path, read_only=True, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")
    for name in wb.sheetnames:
        ws = wb[name]
        rows = ws.iter_rows(max_row=5, values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            print(f"  [{name}] (empty)")
            continue
        headers = [_normalize_header(c) for c in header_row]
        print(f"  [{name}] columns ({len(headers)}): {headers}")
        for i, r in enumerate(rows, start=2):
            if r is None or all(c is None or str(c).strip() == "" for c in r):
                continue
            preview = dict(zip(headers, [str(c)[:40] if c is not None else "" for c in r]))
            print(f"    sample row {i}: {preview}")
    wb.close()


def iter_data_rows(path: Path) -> list[tuple[str, int, dict[str, Any]]]:
    """Yield (sheet_name, excel_row_number, row_dict) for non-empty data rows."""
    wb = load_workbook(path, read_only=True, data_only=True)
    out: list[tuple[str, int, dict[str, Any]]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            continue
        headers = [_normalize_header(c) for c in header_row]
        for idx, row in enumerate(rows_iter, start=2):
            row_dict = _row_to_dict(headers, row)
            if _is_empty_row(row_dict):
                continue
            out.append((sheet_name, idx, row_dict))
    wb.close()
    return out


def _get_supabase() -> Client:
    load_dotenv(PROJECT_ROOT / ".env")
    url = os.environ.get("SUPABASE_URL", "").strip()
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    if not url or not key:
        print(
            "ERROR: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set "
            "(e.g. in .env). Do not hardcode secrets.",
            file=sys.stderr,
        )
        sys.exit(1)
    return create_client(url, key)


def _source_payload(
    slug: str,
    name: str,
    kind: str,
    intelligence_layer: str,
    base_url: str | None,
    owner_team: str | None,
    license_notes: str | None,
    default_refresh: str | None,
    metadata: dict[str, Any],
) -> dict[str, Any]:
    return {
        "slug": slug,
        "name": name[:500],
        "kind": kind[:100],
        "intelligence_layer": intelligence_layer,
        "base_url": base_url,
        "owner_team": owner_team,
        "license_notes": license_notes,
        "default_refresh_frequency": default_refresh,
        "metadata": metadata,
        "is_active": True,
    }


def _fetch_source_by_slug(client: Client, slug: str) -> dict[str, Any] | None:
    res = client.table("sources").select("*").eq("slug", slug).limit(1).execute()
    if res.data:
        return res.data[0]
    return None


def _fetch_dataset_by_source_and_external(
    client: Client, source_id: str, external_id: str
) -> dict[str, Any] | None:
    res = (
        client.table("datasets")
        .select("*")
        .eq("source_id", source_id)
        .eq("external_id", external_id)
        .limit(1)
        .execute()
    )
    if res.data:
        return res.data[0]
    return None


def _fetch_version_by_dataset_and_label(
    client: Client, dataset_id: str, version_label: str
) -> dict[str, Any] | None:
    res = (
        client.table("dataset_versions")
        .select("*")
        .eq("dataset_id", dataset_id)
        .eq("version_label", version_label)
        .limit(1)
        .execute()
    )
    if res.data:
        return res.data[0]
    return None


def import_workbook(
    client: Client,
    path: Path,
    intelligence_layer: str,
    workbook_key: str,
    stats: ImportStats,
) -> None:
    print(f"\n=== Importing {path.name} ({intelligence_layer}) ===")
    if not path.is_file():
        print(f"ERROR: workbook not found: {path}")
        sys.exit(1)

    rows = iter_data_rows(path)
    print(f"  Data rows to process: {len(rows)}")

    slugs = [_stable_slug(intelligence_layer, sh, rd) for sh, _, rd in rows]
    pre_existing: set[str] = set()
    for slug in slugs:
        if _fetch_source_by_slug(client, slug):
            pre_existing.add(slug)
    seen_slugs_in_run: set[str] = set()

    for sheet_name, row_num, row_dict in rows:
        if _is_empty_row(row_dict):
            stats.rows_skipped += 1
            continue

        row_norm = {str(k).lower().strip(): v for k, v in row_dict.items() if k}
        name = _pick_value(row_norm, NAME_KEYS) or f"Unnamed source ({sheet_name} row {row_num})"
        kind = (_pick_value(row_norm, KIND_KEYS) or "file").strip() or "file"
        if len(kind) < 1:
            kind = "file"
        base_url = _pick_value(row_norm, URL_KEYS)
        owner = _pick_value(row_norm, OWNER_KEYS)
        lic = _pick_value(row_norm, LICENSE_KEYS)
        refresh = _pick_value(row_norm, REFRESH_KEYS)
        slug = _stable_slug(intelligence_layer, sheet_name, row_dict)

        meta = {
            "source_map_workbook": workbook_key,
            "sheet": sheet_name,
            "excel_row": row_num,
            "intelligence_layer": intelligence_layer,
            "raw_row": row_dict,
        }

        existed = slug in pre_existing or slug in seen_slugs_in_run
        payload = _source_payload(
            slug=slug,
            name=name,
            kind=kind,
            intelligence_layer=intelligence_layer,
            base_url=base_url,
            owner_team=owner,
            license_notes=lic,
            default_refresh=refresh,
            metadata=meta,
        )

        client.table("sources").upsert(payload, on_conflict="slug").execute()
        if existed:
            stats.sources_updated += 1
        else:
            stats.sources_inserted += 1
        seen_slugs_in_run.add(slug)

        src = _fetch_source_by_slug(client, slug)
        if not src or not src.get("id"):
            print(f"  WARN: could not re-load source slug={slug}")
            continue
        source_id = str(src["id"])

        # One companion dataset per source (tracking); external_id ties back to source slug.
        ds_external = f"{slug}::primary"
        ds_title = name[:500]
        ds_meta = {
            "source_map_workbook": workbook_key,
            "sheet": sheet_name,
            "excel_row": row_num,
            "raw_row": row_dict,
            "linked_source_slug": slug,
        }
        existing_ds = _fetch_dataset_by_source_and_external(client, source_id, ds_external)
        ds_payload = {
            "source_id": source_id,
            "external_id": ds_external,
            "title": ds_title,
            "description": None,
            "access_method": kind,
            "metadata": ds_meta,
        }
        if existing_ds:
            client.table("datasets").update(
                {
                    "title": ds_payload["title"],
                    "access_method": ds_payload["access_method"],
                    "metadata": ds_meta,
                }
            ).eq("id", existing_ds["id"]).execute()
            stats.datasets_updated += 1
            dataset_id = str(existing_ds["id"])
        else:
            ins = client.table("datasets").insert(ds_payload).select("*").execute()
            if not ins.data:
                print(f"  WARN: dataset insert returned no data for source {slug}")
                continue
            dataset_id = str(ins.data[0]["id"])
            stats.datasets_inserted += 1

        ver_meta = {
            "source_map_workbook": workbook_key,
            "sheet": sheet_name,
            "excel_row": row_num,
            "note": "Placeholder version from source-map import; no raw file ingested yet.",
            "raw_row": row_dict,
        }
        existing_ver = _fetch_version_by_dataset_and_label(
            client, dataset_id, DATASET_VERSION_LABEL
        )
        ver_payload = {
            "dataset_id": dataset_id,
            "version_label": DATASET_VERSION_LABEL,
            "ingestion_status": "pending",
            "metadata": ver_meta,
        }
        if existing_ver:
            client.table("dataset_versions").update(
                {"metadata": ver_meta, "ingestion_status": "pending"}
            ).eq("id", existing_ver["id"]).execute()
            stats.versions_updated += 1
        else:
            client.table("dataset_versions").insert(ver_payload).execute()
            stats.versions_inserted += 1


def main() -> int:
    print("Norwegian Career Intelligence — source map import (Spor 1 & Spor 2)")
    print(f"Project root: {PROJECT_ROOT}")

    inspect_workbook(SPOR1_PATH)
    inspect_workbook(SPOR2_PATH)

    client = _get_supabase()
    stats = ImportStats()

    import_workbook(
        client,
        SPOR1_PATH,
        INTELLIGENCE_LAYER_SPOR1,
        "Spor1_Kildekartlegging.xlsx",
        stats,
    )
    import_workbook(
        client,
        SPOR2_PATH,
        INTELLIGENCE_LAYER_SPOR2,
        "Spor2_Kildekartlegging.xlsx",
        stats,
    )

    print("\n--- Summary ---")
    print(f"  sources inserted: {stats.sources_inserted}")
    print(f"  sources updated (upsert): {stats.sources_updated}")
    print(f"  datasets inserted: {stats.datasets_inserted}")
    print(f"  datasets updated: {stats.datasets_updated}")
    print(f"  dataset_versions inserted: {stats.versions_inserted}")
    print(f"  dataset_versions updated: {stats.versions_updated}")
    print(f"  rows skipped (empty): {stats.rows_skipped}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
